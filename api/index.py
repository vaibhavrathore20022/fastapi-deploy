from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import HTMLResponse, StreamingResponse
from enum import Enum
import pandas as pd
from io import BytesIO
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

app = FastAPI()

# Define regions
class RegionEnum(str, Enum):
    INDORE = "INDORE"
    NARMADAPURAM = "NARMADAPURAM"
    REWA = "REWA"
    BHOPALCENTRAL = "BHOPALCENTRAL"
    GWALIOR = "GWALIOR"
    JABALPUR = "JABALPUR"

REGIONS = [region.value for region in RegionEnum]

@app.get("/", response_class=HTMLResponse)
async def form_page():
    region_options = "\n".join([f'<option value="{r}">{r}</option>' for r in REGIONS])
    return f"""
    <html>
        <head><title>Excel Upload</title></head>
        <body>
            <h2>Upload Excel File</h2>
            <form action="/process/" enctype="multipart/form-data" method="post">
                <label>Select Region:</label>
                <select name="region">{region_options}</select><br><br>
                <label>Choose Excel File:</label>
                <input type="file" name="file" accept=".xlsx" required/><br><br>
                <button type="submit">Upload</button>
            </form>
        </body>
    </html>
    """

@app.post("/process/")
async def process_file(region: str = Form(...), file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        return {"error": "Invalid file type"}

    # Extract date range from filename
    match = re.search(r"FROM (\d{1,2}-\w+) TO (\d{1,2}-\w+)", file.filename)
    date_range = f"{match.group(1)} TO {match.group(2)}" if match else "Unknown Date Range"

    df = pd.read_excel(file.file, sheet_name="Reference User Creations")

    df = df[df["STATE"] == "MADHYA PRADESH"]
    df = df[df["REGION_NAME"] == region]

    # Select and rename relevant columns
    columns = [
        "USER_ID", "USER_FIRSTNAME", "CIRCLE_OFFICE", "REGION_NAME", "STATE",
        "APY", "EKYC", "PMSBY", "PMJJBY", "FD", "RD", "TOTAL TXN COUNT", "TOTAL TXN AMT"
    ]
    df = df[columns]

    # Fill missing with 0
    for col in ["APY", "EKYC", "PMSBY", "PMJJBY", "FD", "RD", "TOTAL TXN COUNT"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    # Calculate targets
    total_days = int(match.group(2).split("-")[0]) - int(match.group(1).split("-")[0]) + 1 if match else 30
    targets = {
        "APY": 2, "EKYC": 10, "PMSBY": 10, "PMJJBY": 5,
        "FD": 2, "RD": 2, "TOTAL TXN COUNT": 300
    }

    for key in targets:
        daily = targets[key] / 30
        df[f"{key}_TARGET"] = (daily * total_days)

    def check(val, target): return val >= target

    achieved = []
    for i, row in df.iterrows():
        row_status = []
        for k in targets:
            ok = check(row[k], row[f"{k}_TARGET"])
            row_status.append(ok)
        achieved.append("Achieved" if all(row_status) else "Not Achieved")

    df["STATUS"] = achieved

    inactive_df = df[(df[["APY", "EKYC", "PMSBY", "PMJJBY", "FD", "RD", "TOTAL TXN COUNT"]] == 0).all(axis=1)]

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.drop(columns=[col + "_TARGET" for col in targets]).to_excel(writer, sheet_name="Target Check", index=False)
        inactive_df.drop(columns=[col + "_TARGET" for col in targets]).to_excel(writer, sheet_name="Inactive", index=False)

        wb = writer.book
        ws = wb["Target Check"]

        green_fill = PatternFill(start_color="C6EFCE", fill_type="solid")
        red_font = Font(color="9C0006")
        center = Alignment(horizontal="center")
        bold = Font(bold=True)
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = center
                cell.border = border
                if cell.column_letter in "FGHIJKL":  # KPI columns
                    header = ws.cell(row=1, column=cell.column).value
                    target = df.loc[cell.row - 2, f"{header}_TARGET"]
                    value = df.loc[cell.row - 2, header]
                    if value >= target:
                        cell.fill = green_fill
                    else:
                        cell.font = red_font

        # Header formatting
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="FFD966", fill_type="solid")
            cell.font = bold
            cell.alignment = center
            cell.border = border

        # Add count summary
        ws.cell(row=ws.max_row + 2, column=1).value = "Achieved Count"
        ws.cell(row=ws.max_row, column=2).value = achieved.count("Achieved")
        ws.cell(row=ws.max_row + 1, column=1).value = "Not Achieved Count"
        ws.cell(row=ws.max_row + 1, column=2).value = achieved.count("Not Achieved")

    output.seek(0)
    filename = f"{region}_Processed_{date_range.replace(' ', '_')}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment;filename={filename}"})
